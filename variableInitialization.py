import os
from math import *
from sklearn import preprocessing
import numpy as np
import openpyxl

folderPath = ""

P, Q = 10, 8  # global variables for no. of patients and edge servers
S = 6  # initialize the number of sensors
compare = 0  # set this to 1 to compare performance of all algorithms
patient = []  # list of patients 1, 2, 3, ..., P
healthParams = ['Blood Glucose Level', 'Diastolic Blood Pressure', 'Systolic Blood Pressure', 'Heart Rate', 'Body Temperature',
                'Blood O2 Saturation']  # health parameters recorded by each sensor
normalLowerLimit = [70, 60, 90, 60, 97, 95]  # normal lower limit of health parameters
normalUpperLimit = [100, 80, 120, 100, 99, 100]  # normal upper limit of health parameters
dominance = [3, 4, 5, 6, 1, 2]  # significance of health parameters
recordedSensorValue = []  # health parameters' value recorded by each sensor
healthSeverityIndex = []  # health severity index G of the recorded value
criticalityIndex = []  # criticality index I of the recorded value
criticality = []  # criticality of P patients
EHRSize = []  # [1 - 3 MB] Data size
localCPUUtilization = []  # [0.8 - 1.5 Gcycles/bit] CPU cycles/bit
localComputationalCapacity = []  # [1 - 3 GHz] CPU cycles/sec
localEncryptionCPUUtilization = []  # [0.2 - 0.8 GHz] CPU cycles/bit
localEncryptionEnergy = []  # [1000 - 2000] mAh
localProcessingTime = []  # time required to process EHR on local device (in sec)
transmissionRate = []  # transmission rate between local device for patient p and edge server q
transmissionPower = []  # [0.1 - 20 mW] transmission power between patient p and edge server q
localProcessingMemory = []  # [2 - 6 MB] memory utilization on local device
localProcessingEnergy = []  # [1000 - 2000 mAh] energy consumption on local device
edgeCPUUtilization = []  # [0.2 - 0.7 Gcycles/bit] CPU cycles/bit
edgeComputationalCapacity = []  # [5 - 15 GHz] CPU cycles/sec
offloadingTime = []  # total offloading time (in sec)
offloadingEnergy = []  # total offloading energy usage (in mAh)
offloadingMemory = []  # [2 - 6 MB] memory usage while offloading EHR
maxTransmissionPower = 0.02  # in W
maxEdgeComputationalCapacity = 15 * 1e9  # in Hz
minTransmissionRate = 0.5 * 1e6  # bit/sec
maxLatency = 5  # in sec
maxMemory = 2 * 8 * 1e9  # in bits
totalProcessingTime = [0 for _ in range(P)]  # (1-x)*T_l + x*T_o
totalEnergyConsumption = [0 for _ in range(P)]  # (1-x)*E_l + x*E_o
totalMemoryUsage = [0 for _ in range(P)]  # (1-x)*M_l + x*M_o

V_l = 10  # lower limit of number of validators
V_u = 100  # upper limit of number of validators
N_l = 10  # lower limit of number of transactions per block
N_u = 50  # upper limit of number of transactions per block
downlinkTransmissionRate = 1.2 * 1e6  # bits/sec
uplinkTransmissionRate = 1.3 * 1e6  # bits/sec
requiredComputationalResources = []  # [50 - 100] required computational resources for block verification task for patient p
verificationFeedbackSize = []  # [0.1 - 0.5 Mb] block verification feedback size for patient p
transactionSize = []  # [0.1 - 0.5 kb] transaction size (in bits) of the processed EHR of patient p
resourcesWithValidator = []  # [1 - 100] available computational resources with validator v
computationalCostIncurred = []  # [3 - 300] cost incurred by validator v for block verification of patient p
paymentToCFP = []  # [2 - 200] payment made by validator v to CFP for acquiring resources
maxBlockLatency = 5  # maximum desirable latency (in sec)
maxSecurity = 1e10  # maximum security level
maxCost = 120  # maximum computational cost
mu = 1e14  # weight in penalty function
it = 4  # indicator factor representing the network scale
zeta = 1  # system defined coefficient for varying security level


def initialize():
    global criticalityIndex, recordedSensorValue, criticality, patient
    dataset = openpyxl.load_workbook("./Dataset for People for their Blood Glucose Level with their Superficial body feature readings..xlsx")  # read the dataset
    df = dataset.active  # activate the dataset object
    for p in range(min(P, df.max_row)):  # iterate over the dataset
        criticalityIndex = [0 for _ in range(S)]
        recordedSensorValue = [0.0 for _ in range(S)]
        i = 0
        for s in df.iter_cols(2, df.max_column - 3):
            recordedSensorValue[i] = round(float(s[p + 3].value), 2)
            G = abs((normalUpperLimit[i] - recordedSensorValue[i]) ** 2 - (recordedSensorValue[i] - normalLowerLimit[i]) ** 2) / (
                    abs(normalUpperLimit[i]) + abs(normalLowerLimit[i])) ** 2
            healthSeverityIndex.append(G)
            cIndex = dominance[i] * G
            criticalityIndex[i] = cIndex
            i += 1
        criticalityIndex = list(
            preprocessing.MinMaxScaler(feature_range=(0, 1)).fit_transform(np.array(criticalityIndex).reshape(-1, 1)).reshape(-1))  # min-max normalization of criticality index
        C = sum(criticalityIndex) / S  # calculate criticality of patient 'p'
        criticality.append(C)
        patient.append(p)
        EHRSize.append(np.random.uniform(1 * 8 * 1e6, 3 * 8 * 1e6))
        localCPUUtilization.append(np.random.uniform(0.8 * 1e9, 1.5 * 1e9))
        localComputationalCapacity.append(np.random.uniform(1 * 1e9, 3 * 1e9))
        localEncryptionCPUUtilization.append(np.random.uniform(0.2 * 1e9, 0.8 * 1e9))
        localEncryptionEnergy.append(np.random.randint(1000, 2000))
        localProcessingEnergy.append(np.random.randint(1000, 2000))
        localProcessingMemory.append(np.random.uniform(2 * 8 * 1e6, 6 * 8 * 1e6))
        T_l = (EHRSize[p] * localCPUUtilization[p]) / localComputationalCapacity[p]  # local processing time
        localProcessingTime.append(T_l)
        trRate = []
        trPower = []
        offTime = []
        offEnergy = []
        for q in range(Q):
            edgeCPUUtilization.append(np.random.uniform(0.2 * 1e9, 0.7 * 1e9))
            edgeComputationalCapacity.append(np.random.uniform(5 * 1e9, 15 * 1e9))
            SINR = np.random.randint(13, 20)  # [13 - 20] in dB
            Z = np.random.randint(5, 15)  # [5 - 15] PRB
            Y = np.random.randint(5 * 1e6, 15 * 1e6)  # [5 - 15 MHz] channel bandwidth
            rate = Y * Z * log(1 + SINR, 2)  # transmission rate
            trRate.append(rate)
            trPower.append(np.random.uniform(0.1 / 1e3, 20 / 1e3))
            E_pq = (trPower[q] * EHRSize[p]) / trRate[q]  # transmission energy
            T_po = (EHRSize[p] * localEncryptionEnergy[p]) / localComputationalCapacity[p] + (
                    EHRSize[p] * edgeCPUUtilization[q]) / edgeComputationalCapacity[q] + EHRSize[p] / trRate[q]
            E_po = localEncryptionEnergy[p] + E_pq
            offTime.append(T_po)
            offEnergy.append(E_po)
        transmissionRate.append(trRate)
        transmissionPower.append(trPower)
        offloadingTime.append(offTime)
        offloadingEnergy.append(offEnergy)
        offloadingMemory.append(np.random.uniform(2 * 8 * 1e6, 6 * 8 * 1e6))
        # initialization of blockchain utility objective's variables begins here
        requiredComputationalResources.append(np.random.randint(50, 100))
        verificationFeedbackSize.append(np.random.uniform(0.1 * 1e6, 0.5 * 1e6))
        transactionSize.append(np.random.uniform(0.1 * 1e3, 0.5 * 1e3))
        compCost = []
        for v in range(V_u + 1):
            compCost.append(np.random.uniform(3, 300))
        computationalCostIncurred.append(compCost)
    for v in range(V_u + 1):
        resourcesWithValidator.append(np.random.randint(1, 100))
        paymentToCFP.append(np.random.uniform(2, 200))
    patient = np.array(patient)[(-np.array(criticality)).argsort()]
    # Here onwards we decide the path of folder where comparison graphs should be saved
    if compare:
        baseFolder = "Graphical outcomes 1"  # by default the graphs are saved in base folder "Graphical outcomes 1"
        basePath = "./Graphs/"  # the base path is "./Graphs"
        if not os.path.exists("./Graphs"):  # if the base path does not exist, create it
            os.mkdir("./Graphs")
        global folderPath
        folderPath = os.path.join(basePath + baseFolder)  # set the folder path by combining base path and base folder
        i = 1  # set this to 1 to get next available base folder
        if os.path.exists(folderPath):  # if the defined folder path already exists, create a new base folder to save graphs
            while True:
                i += 1  # increment to get next available base folder
                baseFolder = "Graphical outcomes"
                folderName = f"{baseFolder} {i}"  # set the folder name
                folderPath = os.path.join(basePath, folderName)
                if not os.path.exists(folderPath):  # check if the new folder path exists
                    os.mkdir(folderPath)  # if not, then create it and break the loop
                    break
        else:
            os.mkdir(folderPath)  # if the folder path does not exist, create it


# function to get the path of folder where comparison graphs should be saved
def getFolderPath():
    return folderPath


# function to get the list of indices of patients sorted in decreasing order of their criticality
def getPatientList():
    return patient


# Function to generate the whale population
def initializeDataWhalePopulation():
    whale = [[0 for _ in range(Q)] for _ in range(Q + 1)]  # list to store the whale population
    q = 0
    while q < Q:
        whale[q][q] = 1  # offloading happens on qth edge server
        q += 1
    # whale = []
    # for counter in range(2 ** Q):
    #     c = counter
    #     w = []
    #     for q in range(Q):
    #         if c & 1:
    #             w.append(1)
    #         else:
    #             w.append(0)
    #         c >>= 1
    #     whale.append(w)
    return whale


# Function to calculate the fitness of a whale
def calculateDataWhaleFitness(p, whale):
    fitness = []  # initialize an empty list to store fitness of whales representing solutions for data offloading objective
    for w in range(len(whale)):  # for each whale 'w'
        f = 0  # helper variable to find fitness of whales for a patient 'p'
        penalty = 0  # helper variable to find penalty of whale solution for a patient 'p'
        xTo, xTl, xEo, xEl, xMo, xMl = 0, 0, 0, 0, 0, 0  # variables to find constraint-related mathematical terms
        t = (1 - (1 in whale[w])) * localProcessingTime[p]  # helper variables to find time, energy and memory consumed by given whale for patient 'p'
        e = (1 - (1 in whale[w])) * localProcessingEnergy[p]
        m = (1 - (1 in whale[w])) * localProcessingMemory[p]
        for q in range(Q):  # for each edge server
            t += whale[w][q] * offloadingTime[p][q]
            e += whale[w][q] * offloadingEnergy[p][q]
            m += whale[w][q] * offloadingMemory[p]
            if whale[w][q]:  # if offloading happens at edge server 'q' in whale 'w'
                if transmissionRate[p][q] < minTransmissionRate:  # if the constraint on transmission rate is violated
                    penalty += mu * (minTransmissionRate - transmissionRate[p][q]) ** 2  # add to the penalty
                if transmissionPower[p][q] > maxTransmissionPower:  # if the constraint on transmission power is violated
                    penalty += mu * (transmissionPower[p][q] - maxTransmissionPower) ** 2  # add to the penalty
                if transmissionPower[p][q] <= 0:  # if second constraint on transmission power is violoated
                    penalty += mu * (-transmissionPower[p][q]) ** 2  # add to the penalty
                if edgeComputationalCapacity[q] <= 0:  # if constraint on edge computational capacity is violated
                    penalty += mu * (-edgeComputationalCapacity[q]) ** 2  # add to the penalty
                if edgeComputationalCapacity[q] > maxEdgeComputationalCapacity:  # if second constraint on edge computational capacity is violated
                    penalty += mu * (edgeComputationalCapacity[q] - maxEdgeComputationalCapacity) ** 2  # add to the penalty
            xTo += whale[w][q] * offloadingTime[p][q]
            xTl += (1 - whale[w][q]) * localProcessingTime[p]
            xEo += whale[w][q] * offloadingEnergy[p][q]
            xEl += (1 - whale[w][q]) * localProcessingEnergy[p]
            xMo += whale[w][q] * offloadingMemory[p]
            xMl += (1 - whale[w][q]) * localProcessingMemory[p]
        totalProcessingTime[p] = t
        totalEnergyConsumption[p] = e
        totalMemoryUsage[p] = m
        if xTo > xTl:  # if total processing time on edge server is greater than that on local device, the constraint is violated
            penalty += mu * (xTo - xTl) ** 2  # add to the penalty
        if xMo > xMl:  # if total processing memory on edge server is greater than that on local device, the constraint is violated
            penalty += mu * (xMo - xMl) ** 2  # add to the penalty
        if xEo > xEl:  # if total processing energy on edge server is greater than that on local device, the constraint is violated
            penalty += mu * (xEo - xEl) ** 2  # add to the penalty
        f += totalProcessingTime[p] * criticality[p] + totalEnergyConsumption[p] * (criticality[p] ** 2) + totalMemoryUsage[p] * (
                criticality[p] ** 2)  # add to fitness the data offloading objective's value
        if totalProcessingTime[p] > maxLatency:  # if total processing time is greater than maximum desirable latency, the constraint is volated
            penalty += mu * (totalProcessingTime[p] - maxLatency) ** 2  # add to the penalty
        if totalMemoryUsage[p] > maxMemory:  # if total memory consumed is greater than maximum available memory, the constraint is violated
            penalty += mu * (totalMemoryUsage[p] - maxMemory) ** 2  # add to the penalty
        fitness.append(f + penalty)  # get the fitness of whale 'w'
    return fitness


# Function to generate the whale population
def initializeBlockchainWhalePopulation():
    whale = []  # initialize empty list to store whale population
    for n in range(N_l, N_u + 1):  # for each number from the lowest to the highest number of transactions per block
        for v in range(V_l, V_u + 1):  # for each number from the lowest to the highest number of validators
            whale.append([n, v])  # create a whale
    return whale


# Function to calculate the fitness of a whale
def calculateBlockchainWhaleFitness(whale):
    fitness = []  # initialize an empty list to store fitness of each whale
    for w in range(len(whale)):  # for each whale in population
        f = 0  # helper variable to find sum of utilities for each patient for a given whale
        penalty = 0  # helper variable to find sum of penalties involving the constraints satisfied or violated by a whale
        for p in range(P):  # for each patient
            f += calculateUtility(whale[w][0], whale[w][1], p)  # add to fitness the utility of solution given by whale 'w'
            for v in range(whale[w][1]):  # for all validators presumed by whale 'w'
                if computationalCostIncurred[p][v] < paymentToCFP[v] * resourcesWithValidator[v]:  # if this is true, the constraint on computational cost is violated
                    penalty += mu * (paymentToCFP[v] * resourcesWithValidator[v] - computationalCostIncurred[p][v]) ** 2  # add to the penalty
        if whale[w][1] > V_u:  # if constraint on upper limit of validators is violated
            penalty += mu * (whale[w][1] - V_u) ** 2  # add to the penalty
        if whale[w][1] < V_l:  # if constraint on lower limit of validators is violated
            penalty += mu * (V_l - whale[w][1]) ** 2  # add to the penalty
        if whale[w][0] > N_u:  # if constraint on upper limit of number of transactions per block is violated
            penalty += mu * (whale[w][1] - N_u) ** 2  # add to the penalty
        if whale[w][0] < N_l:  # if constraint on lower limit of number of transactions per block is violated
            penalty += mu * (N_l - whale[w][0]) ** 2  # add to the penalty
        fitness.append(f + penalty)  # get the fitness of whale 'w'
    return fitness


# function to calculate blockchain utility for given values of number of transactions per second (N), number of validators (V) and patient's index (p)
def calculateUtility(N, V, p):
    latency = (N * transactionSize[p]) / downlinkTransmissionRate + requiredComputationalResources[p] / min(resourcesWithValidator) + N * transactionSize[p] * V + \
              verificationFeedbackSize[p] / uplinkTransmissionRate  # get the latency involved in block addition
    security = zeta * V ** it  # get the security level maintained while block validation
    cost = sum(computationalCostIncurred[p]) / N  # get the cost incurred by validators for verifying block's data
    utility = (latency * criticality[p]) / maxBlockLatency + (maxSecurity * (criticality[p] ** 2)) / security + (cost * (criticality[p] ** 2)) / maxCost  # calculate utility
    return utility


# Utility function to print a matrix
def printMatrix(mat):
    for i in range(len(mat)):  # iterate over rows
        for j in range(len(mat[i])):  # iterate over columns
            print(mat[i][j], end=' ')  # print the element at ith row and jth column
        print()  # print newline after each row
