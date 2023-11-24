import hashlib
import base64
import json
import os
import random
import shutil
import openpyxl
from collections import Counter
from datetime import datetime
from cryptography.hazmat.primitives import serialization, hashes
from cryptography.hazmat.primitives.asymmetric import rsa, padding
from cryptography.fernet import Fernet
import socket
import threading
from variableInitialization import requiredComputationalResources, P, healthParams
from ACO_Blockchain_Optimization import export_ACO_Blockchain_Metrics
from ACO_Data_Offloading import export_ACO_Data_Metrics

nodes, acoDataSolution, acoBlockchainSolution, faulty_nodes = [], [[]], [], []
node_public_key, node_private_key = {}, {}
re = 0
blockchain = None


# class to represent a block of blockchain
class Block:
    def __init__(self, index, timestamp, data, previous_hash, patient_id, transaction_initiator):
        self.index = index
        self.timestamp = timestamp
        self.data = data
        self.previous_hash = previous_hash
        self.patient_id = patient_id
        self.transaction_initiator = transaction_initiator
        self.hash = self.calculate_hash()

    # function to calculate hash value of the block using SHA256 algorithm
    def calculate_hash(self):
        hash_data = (
                str(self.index)
                + str(self.timestamp)
                + str(self.data)
                + str(self.previous_hash)
                + str(self.patient_id)
                + str(self.transaction_initiator)
        )
        return hashlib.sha256(hash_data.encode()).hexdigest()


# class to represent a node (validator)
class Node:
    def __init__(self, name, public_key, private_key, resources, node_socket):
        self.name = name
        self.public_key = public_key
        self.private_key = private_key
        self.resources = resources
        self.node_socket = node_socket


# class to represent a blockchain
class Blockchain:
    def __init__(self):
        self.chain = []
        self.pending_transactions = []
        self.access_mapping = {}

    # function to create the starting block of blockchain
    def create_genesis_block(self):
        genesis_block = Block(0, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Genesis Block", "0", "", "Hospital")
        self.chain.append(genesis_block)

    # function to get the last block of blockchain
    def get_last_K_blocks(self, K):
        return self.chain[-K:]

    # function to add a block in blockchain
    def add_block(self, new_block):
        if len(self.chain) == 0:  # if the chain is empty, directly add the block
            self.chain.append(new_block)
            return True  # return true if the block is added successfully, else return false
        if new_block.previous_hash == self.get_last_K_blocks(1)[0].hash:  # if the chain is not empty, verify if the hash of new block matches with that of the last block
            self.chain.append(new_block)
            return True
        return False

    # function to keep track of pending transactions (not added as a block in the chain)
    def add_transaction(self, transaction):
        self.pending_transactions.append(transaction)

    # function to maintain list of users with IDs = 'entity_ids' who can access record of patient with ID = 'patient_id'
    def create_access_mapping(self, patient_id, entity_ids):
        self.access_mapping[patient_id] = entity_ids

    # function to validate pending transactions
    def validate_pending_transactions(self):
        while len(self.pending_transactions) > 0:
            previous_hash = self.get_last_K_blocks(1)[0].hash  # get the hash of previous block
            patient_id = self.pending_transactions[0][0]  # get the patient ID of patient whose transaction is pending
            data, transaction_to_remove = [], []
            # Extract the N* transactions of processed EHRs for the obtained patient ID. If there are less than N* such transactions, extract all
            for pt in range(len(self.pending_transactions)):
                if len(data) == acoBlockchainSolution[0]:
                    break
                if self.pending_transactions[pt][0] == patient_id:
                    self.pending_transactions[pt].pop(0)
                    data.append(self.pending_transactions[pt])
                    transaction_to_remove.append(pt)  # list to store transactions to be completed
            # Remove N* transactions with the obtained patient ID from the list of pending transactions
            self.pending_transactions = [transaction for index, transaction in enumerate(self.pending_transactions) if index not in transaction_to_remove]
            lis = [ind + 1 for ind, value in enumerate(acoDataSolution[patient_id - 1]) if value == 1]  # get the index of edge server on which the data of patient with ID = 'patient_id' is uploaded
            transaction_initiator = f"Edge server: {lis}" if len(lis) > 0 else "Local device: "  # identify the source of processed EHR
            try:
                new_block = Block(
                    len(self.chain),
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    encrypt_EHR(len(self.chain), str(data)),
                    previous_hash,
                    patient_id,
                    transaction_initiator
                )  # create new block for this transaction
                # Save the JSON data to a file
                filename = "new_block.json"
                with open(filename, "w") as file:
                    json.dump(new_block.__dict__, file)
                os.chmod(filename, 0o444)
                broadcast_block(filename)  # Broadcast the block to all connected nodes
            except (OSError, Exception):
                print("Error while creating the block.")
                return
            print("Awaiting validators' approval...")
            valid_block = True  # set this to true
            block_acceptance_message = "Enter 1 to accept or 0 to reject the block: "
            broadcast_message(block_acceptance_message)
            responses = []
            for node in nodes:
                response = node.node_socket.recv(2 ** 15).decode()
                responses.append(response)
            if not all(response == "1" for response in responses):
                valid_block = False
            if valid_block:  # if the block is valid
                self.add_block(new_block)  # add it to blockchain
                broadcast_message("Block accepted and added to the blockchain.")
                print("Block accepted and added to the blockchain.")  # display block acceptance message
                self.store_block(True, new_block)  # store the copy of this block with each node
            else:
                broadcast_message("Block rejected by one or more validators.")
                print("Block rejected by one or more validators.")  # display block rejection message
                self.store_block(False, new_block)  # pass False to store_block function

    # function to store copy of block with each node
    def store_block(self, accepted, new_block):
        nodes_with_lack_of_resources = []
        for node in nodes:
            folder_path = "./blockchain/{}".format(node.name)  # get the location of each node
            os.makedirs(folder_path, exist_ok=True)  # create the location of node if it does not exist
            file_name = f"block {len(self.chain) - 1}.json"  # define the name of the block to be stored
            file_path = os.path.join(folder_path, file_name)  # combine the node's location and the block's name
            if accepted:  # if the block is valid
                with open(file_path, "w") as file:  # create the block's file
                    json.dump(new_block.__dict__, file)
            if node.resources > requiredComputationalResources[new_block.patient_id - 1]:  # if the node has more resources than required for validation of transaction of a patient
                node.resources -= requiredComputationalResources[new_block.patient_id - 1]
            else:
                nodes_with_lack_of_resources.append(node.name.split()[1])  # otherwise the node purchases the required resources from cloud/fog provider
                node.resources = 0  # node's resources again become zero, since only the required amount of resources was purchased
        if len(nodes_with_lack_of_resources) > 0:
            print("Nodes", ", ".join(nodes_with_lack_of_resources), "purchase more resources from CFP.")

    # function to check the validity of data access of a patient with ID = 'patient_id' by a user with ID = 'entity_id'
    def check_access(self, patient_id, entity_id):
        if patient_id in self.access_mapping and entity_id in self.access_mapping[patient_id]:  # if patient_id exists and the patient has allowed 'entity_id' to access their data
            return True
        return False

    # function to fetch the block with a particular patient's ID from the blockchain
    def get_blocks_by_patient_id(self, patient_id):
        blocks = []
        for block in self.chain:  # for each block in chain
            if block.patient_id == patient_id:  # if a match of patient_id is found
                blocks.append(block)
        return blocks


# function to make a directory deletable
def make_all_deletable(directory):
    try:
        for root, dirs, files in os.walk(directory):
            # Change permissions for directories
            for dir_name in dirs:
                dir_path = os.path.join(root, dir_name)
                os.chmod(dir_path, 0o644)  # Allow read-write permissions for owner
            # Change permissions for files
            for file_name in files:
                file_path = os.path.join(root, file_name)
                os.chmod(file_path, 0o644)  # Allow read-write permissions for owner
    except Exception as e:
        print(f"An error occurred: {str(e)}")


# function to re-initialize blockchain in case of system failure
def reinitialize_blockchain(blockchain):
    if not os.path.exists("./blockchain/Node 1"):  # if the blockchain does not previously exist, no re-initializing is needed
        return False  # return false to indicate that the blockchain is not re-created
    print("Re-starting blockchain...")
    max_blocks_node = find_max_blocks_node()  # get the node index with maximum no. of blocks
    file_names = sorted(os.listdir(f"./blockchain/Node {max_blocks_node}"), key=lambda x: int(x.split()[1].split('.')[0]))
    for file_name in file_names:  # for each block with max_blocks_node
        if os.path.isfile(f"./blockchain/Node {max_blocks_node}/{file_name}"):  # check if that block still exists with this node to prevent integrity violations
            with open(f"./blockchain/Node {max_blocks_node}/{file_name}", "r") as file:
                try:
                    block_data = json.load(file)  # load this block's data
                except (json.decoder.JSONDecodeError, Exception):
                    block_data = {}
            index = block_data["index"] if "index" in block_data else None  # get the index of the block
            timestamp = block_data["timestamp"] if "timestamp" in block_data else None  # get the timestamp of the block
            data = block_data["data"] if "data" in block_data else None  # get the patient's data stored in block
            previous_hash = block_data["previous_hash"] if "previous_hash" in block_data else None  # get the previous hash value of the block
            patient_id = block_data["patient_id"] if "patient_id" in block_data else None  # get the patient_id
            transaction_initiator = block_data["transaction_initiator"] if "transaction_initiator" in block_data else None  # get the transaction initiator
            hash = block_data["hash"] if "hash" in block_data else None  # get the hash value of the block
            new_block = Block(index, timestamp, data, previous_hash, patient_id, transaction_initiator)  # create the block
            new_block.hash = hash  # set its hash value
            blockchain.add_block(new_block)  # add the block to blockchain
    if os.path.isfile("./blockchain/access_mapping.json"):  # if the file of access mapping is found
        with open("./blockchain/access_mapping.json", "r") as file:
            try:
                access_mapping_restored = json.load(file)  # load the mapping
            except (json.decoder.JSONDecodeError, Exception):
                file.close()
                print("Access mapping could not be restored.")
                make_all_deletable("./blockchain")
                shutil.rmtree("./blockchain")
                return False
        for patient_id, entity_ids in access_mapping_restored.items():  # restore the access mapping
            try:
                patient_id = int(patient_id)
                entity_ids = [int(entity_id) for entity_id in entity_ids]
                blockchain.create_access_mapping(patient_id, entity_ids)  # create new mapping for each item in file
            except (ValueError, Exception):
                print("Access mapping is corrupted.")
                make_all_deletable("./blockchain")
                shutil.rmtree("./blockchain")
                return False
    return True  # return true to indicate that the blockchain is re-created


# function to check and prevent the integrity violations in blockchain for data security
def check_blockchain_integrity(blockchain):
    total_blocks = len(blockchain.chain)  # get the total no. of blocks in the chain
    total_nodes = acoBlockchainSolution[1]  # get the total no. of validators regarded as optimal by the ACO solution
    block_counts = Counter()  # create a counter variable to trace the no. of copies of a block
    majority_count = total_nodes // 2 + 1  # get the threshold for a copy of a block to be deemed valid
    correct_data = [0.0 for _ in range(total_blocks)]  # create a list to store the correct data for each block
    extracted_data = [[0.0 for _ in range(total_nodes)] for __ in range(total_blocks)]  # create a list to store the data of each block held by each node
    for block in range(total_blocks):  # for each block
        for node in range(1, total_nodes + 1):  # for each node
            block_file = f"./blockchain/Node {node}/block {block}.json"  # get the location of the block
            if os.path.isfile(block_file):  # if the block exists at the obtained location
                with open(block_file, "r") as file:
                    try:
                        block_data = json.load(file)  # load the data of block
                    except (json.decoder.JSONDecodeError, Exception):
                        block_data = {}
                block_data = hashlib.sha256(str(block_data).encode()).hexdigest()  # create the hash of the data using SHA256 algorithm
                extracted_data[block][node - 1] = block_data  # store the data of block with current node
                block_counts[block_data] += 1  # increment the count of this block's data
        try:
            most_common_block_data = block_counts.most_common(1)[0]  # get the most frequent copy of data for this block
            max_count = most_common_block_data[1]  # get the frequency of the most frequent copy of data for this block
            frequency = sum(
                count == max_count for _, count in block_counts.items())  # get the sum of frequencies of different data for this block where the frequency is the highest
            # if only one type of copy of data has the highest frequency, and it is greater than or equal to the threshold for data to be valid
            if frequency == 1 and most_common_block_data[1] >= majority_count:
                correct_data[block] = most_common_block_data[0]  # store the correct copy of block's data
            else:
                correct_data[block] = None  # store None
        except (IndexError, Exception):
            correct_data[block] = None
        block_counts.clear()  # clear all data counts for next block's data verification
    global faulty_nodes
    faulty_nodes = [0 for _ in range(acoBlockchainSolution[1])]  # create a list to store malicious/faulty nodes
    for block in range(total_blocks):  # for each block
        for node in range(1, total_nodes + 1):  # for each node
            if extracted_data[block][node - 1] != correct_data[block]:  # if the data of block held by node is not correct
                faulty_nodes[node - 1] = 1  # set this node as faulty
    for node in range(1, total_nodes + 1):  # for each node
        if faulty_nodes[node - 1] == 1:  # if it is faulty, discard its blockchain
            print(f"Discarding the blockchain and resources of node {node} due to block(s) tampering or inconsistency.")
            if os.path.exists(f"./blockchain/Node {node}"):  # if the node's location exists
                for block in os.listdir(f"./blockchain/Node {node}"):  # for each block held by faulty node
                    file_path = os.path.join(f"./blockchain/Node {node}", block)  # combine it with block's file
                    if os.path.isfile(file_path):  # if the block exists
                        os.remove(file_path)  # delete the block at the obtained path
    print("Blockchain integrity checked and faulty nodes (if any) penalized.")
    non_faulty_node = 0  # set this to 0 to get the index of first non-faulty node
    for node in range(1, total_nodes + 1):  # for each node
        if faulty_nodes[node - 1] == 0:  # if it is not faulty
            non_faulty_node = node  # store its index
            break
    if non_faulty_node == 0:  # if no non-faulty node is found
        print("All nodes have invalid blockchain.")
        make_all_deletable("./blockchain")
        shutil.rmtree("./blockchain")
        return False
    else:
        source_dir = f"./blockchain/Node {non_faulty_node}"  # get the location of non-faulty node
        for node in range(1, total_nodes + 1):  # for each node
            if faulty_nodes[node - 1] == 1:  # if node is faulty
                destination_dir = f"./blockchain/Node {node}"  # get the location of faulty node
                if not os.path.exists(destination_dir):  # if the node has newly joined to validate blocks
                    faulty_nodes[node - 1] = 0  # set it as non-faulty
                    os.mkdir(destination_dir)  # create the node's directory
                for file_name in os.listdir(source_dir):  # for each block held by non-faulty node
                    source_path = os.path.join(source_dir, file_name)  # get the location of the block
                    destination_path = os.path.join(destination_dir, file_name)  # get the location of block to be stored with faulty node
                    shutil.copy(source_path, destination_path)  # copy the correct block for faulty node
        if 1 in faulty_nodes:  # if at least 1 faulty node was found
            print("Faulty nodes forced to correct their blockchain.")
        return True  # blockchain integrity has been maintained


# find the node with maximum number of blocks to get the valid blockchain
def find_max_blocks_node():
    i = 1  # set this to 1 for Node 1
    max_blocks = 0  # set maximum no. of block to 0
    max_blocks_node = 1  # set the index of such node to 0
    while os.path.exists(f"./blockchain/Node {i}"):  # if 'Node i' is present
        if len(os.listdir(f"./blockchain/Node {i}")) > max_blocks:  # get the no. of blocks possessed by 'Node i' and update maximum no. of blocks accordingly
            max_blocks = len(os.listdir(f"./blockchain/Node {i}"))
            max_blocks_node = i  # update the index of max blocks' node
        i += 1  # increment to check for next nodes
    return max_blocks_node


# function to encrypt the given EHR for a block
def encrypt_EHR(block_index, EHR):
    # Generate RSA keys
    private_key = rsa.generate_private_key(public_exponent=65537, key_size=8192)
    # Serialize private key to PEM (Privacy Enhanced Mail) format
    private_key_pem = private_key.private_bytes(encoding=serialization.Encoding.PEM, format=serialization.PrivateFormat.PKCS8, encryption_algorithm=serialization.NoEncryption())  # No encryption for now
    # Encrypt the private key using Fernet
    encryption_key = Fernet.generate_key()
    cipher_suite = Fernet(encryption_key)
    encrypted_private_key = cipher_suite.encrypt(private_key_pem)
    # Store the encrypted private key and Fernet key in a JSON file
    data_to_store = {"encrypted_private_key": base64.b64encode(encrypted_private_key).decode(), "key": base64.b64encode(encryption_key).decode()}
    file_path = f"./blockchain/encrypted_keys/encrypted_private_key {block_index}.json"
    # Store the encrypted private key in a file
    os.makedirs("./blockchain/encrypted_keys/", exist_ok=True)
    if os.path.isfile(file_path):
        os.chmod(file_path, 0o644)
    with open(file_path, "w") as file:
        json.dump(data_to_store, file)
    os.chmod(file_path, 0o444)
    # Simulate encrypting data using the public key
    # OAEP = (Optimal Asymmetric Encryption Padding), MGF = (Mask Generation function)
    data_to_encrypt = EHR.encode("utf-8")
    encrypted_data = private_key.public_key().encrypt(data_to_encrypt, padding.OAEP(mgf=padding.MGF1(algorithm=hashes.SHA256()), algorithm=hashes.SHA256(), label=None))
    encrypted_data = base64.b64encode(encrypted_data).decode()  # Encode bytes as base64
    return encrypted_data


# function to decrypt the encrypted EHR for a block
def decrypt_EHR(block_index, encrypted_data):
    # Read the encrypted private key from the file
    file_path = f"./blockchain/encrypted_keys/encrypted_private_key {block_index}.json"
    if os.path.exists(file_path):
        with open(file_path, "r") as file:
            data = json.load(file)
            encrypted_private_key = base64.b64decode(data["encrypted_private_key"])
            fernet_key = base64.b64decode(data["key"])
    else:
        raise Exception
    # Decrypt the encrypted private key using the Fernet key
    cipher_suite = Fernet(fernet_key)
    decrypted_private_key_pem = cipher_suite.decrypt(encrypted_private_key)
    # Load the decrypted private key
    decrypted_private_key = serialization.load_pem_private_key(decrypted_private_key_pem, password=None)  # No password because it's already decrypted
    # Decrypt the data using the private key
    decrypted_data = decrypted_private_key.decrypt(base64.b64decode(encrypted_data.encode()), padding.OAEP(mgf=padding.MGF1(algorithm=hashes.SHA256()), algorithm=hashes.SHA256(), label=None))
    return decrypted_data.decode()


# function to start blockchain
def startBlockchain(data=None):
    global nodes, acoDataSolution, acoBlockchainSolution, re, blockchain
    blockchain = Blockchain()  # create object of class Blockchain
    acoDataSolution = export_ACO_Data_Metrics()[0]  # get the solution of data offloading by ACO
    acoBlockchainSolution = export_ACO_Blockchain_Metrics()[0]  # get the solution of blockchain utility by ACO
    re = reinitialize_blockchain(blockchain)  # re-initialize blockchain if it was stopped due to unforeseen failure(s)
    if re:
        check = check_blockchain_integrity(blockchain)
        if not check:
            re = 0
    if not re:
        print("Creating blockchain...")
        blockchain = Blockchain()
        blockchain.create_genesis_block()
    dataset = openpyxl.load_workbook("./Dataset for People for their Blood Glucose Level with their Superficial body feature readings..xlsx")
    df = dataset.active
    server = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    server.bind(('127.0.0.1', 8081))
    server.listen(acoBlockchainSolution[1])
    print("Server is listening for connections...")
    while len(nodes) < acoBlockchainSolution[1] - re:
        node_socket, address = server.accept()
        node_handler = threading.Thread(target=handle_node, args=(node_socket,))
        node_handler.start()
        print(f"Accepted connection from address: {address}")
    while True:
        transaction = [P + 1]
        while transaction[0] > P:
            transaction = [int(input("Enter patient ID (or -1 to stop adding transactions): "))]
            if transaction[0] > P:
                print(f"Please enter a patient ID between 1 and {P} inclusive.")
            else:
                break
        if transaction[0] == -1:
            break
        if transaction[0] not in blockchain.access_mapping:
            entity_IDs = [int(x) for x in input(f"Enter the user IDs (space-separated) who can access patient {transaction[0]}'s data: ").split()]
            blockchain.create_access_mapping(transaction[0], entity_IDs)
        if data is None:
            i = 0
            for s in df.iter_cols(2, df.max_column - 3):
                transaction.append(healthParams[i] + ": " + str(s[transaction[0] + 2].value))
                i += 1
        else:
            transaction.append(data)
        blockchain.add_transaction(transaction)
        choice_of_more_transactions = input("Want to add more transactions? (y/n): ")
        if str(choice_of_more_transactions).lower() == "y":
            continue
        blockchain.validate_pending_transactions()
        check_blockchain_integrity(blockchain)

    if os.path.isfile("./blockchain/access_mapping.json"):
        os.chmod("./blockchain/access_mapping.json", 0o644)
    with open("./blockchain/access_mapping.json", "w") as file:
        json.dump(blockchain.access_mapping, file)
    os.chmod("./blockchain/access_mapping.json", 0o444)
    os.chmod("./blockchain/private_keys.json", 0o444)
    os.chmod("./blockchain/public_keys.json", 0o444)

    for node in nodes:
        node.node_socket.close()

    # Access patient data
    while True:
        patient_id = P + 1
        while patient_id > P:
            patient_id = int(input("Enter the patient ID to access the data (or -1 to stop accessing): "))
            if patient_id > P:
                print(f"Please enter a patient ID between 1 and {P} inclusive.")
            else:
                break
        if patient_id == -1:
            break
        entity_id = int(input("Enter the user ID: "))

        if blockchain.check_access(patient_id, entity_id):
            blocks = blockchain.get_blocks_by_patient_id(patient_id)
            if len(blocks) > 0:
                for block in blocks:
                    try:
                        EHR = decrypt_EHR(block.index, str(block.data))
                        print("Access granted. Data in block {}:\n{}\nTimestamp: {}".format(block.index, EHR, block.timestamp))
                    except (FileNotFoundError, Exception):
                        print(f"Error retrieving data for patient {patient_id} at timestamp {block.timestamp}")
            else:
                print(f"No block found for the patient ID {patient_id}.")
        else:
            print("Access denied.")


# function to broadcast a message to all connected nodes
def broadcast_message(message):
    global nodes
    for node in nodes:
        node.node_socket.send(message.encode())


# function to broadcast a JSON file to all connected nodes
def broadcast_block(filename):
    global nodes
    for node in nodes:
        try:
            with open(filename, "rb") as file:
                file_data = file.read()
                node.node_socket.send(file_data)
        except Exception as e:
            print(f"Error broadcasting JSON file to nodes: {str(e)}")
    os.chmod(filename, 0o644)
    os.remove(filename)


# function to handle each connected node
def handle_node(node_socket):
    global nodes, node_public_key, node_private_key, re, blockchain
    flag = 0
    i = len(nodes)
    name = f"Node {i + 1}"
    if not re:
        public_key = hashlib.sha256(name.encode()).hexdigest()
        private_key = hashlib.sha256(public_key.encode()).hexdigest()
    else:
        if os.path.isfile("./blockchain/public_keys.json") and os.path.isfile("./blockchain/private_keys.json"):
            with open("./blockchain/public_keys.json", "r") as file:
                try:
                    public_keys_restored = json.load(file)
                except (json.decoder.JSONDecodeError, Exception):
                    public_keys_restored = {}
            with open("./blockchain/private_keys.json", "r") as file:
                try:
                    private_keys_restored = json.load(file)
                except (json.decoder.JSONDecodeError, Exception):
                    private_keys_restored = {}
            if i + 1 in public_keys_restored and i + 1 in private_keys_restored:
                public_key = public_keys_restored[i + 1]
                private_key = private_keys_restored[i + 1]
            else:
                flag = 1
                public_key = hashlib.sha256(name.encode()).hexdigest()
                private_key = hashlib.sha256(public_key.encode()).hexdigest()
                destination_dir = "./blockchain/{}".format(name)
                os.makedirs(destination_dir, exist_ok=True)
                max_blocks_node = find_max_blocks_node()
                source_dir = f"./blockchain/Node {max_blocks_node}"
                if source_dir != destination_dir:
                    for file_name in os.listdir(source_dir):
                        source_path = os.path.join(source_dir, file_name)
                        destination_path = os.path.join(destination_dir, file_name)
                        shutil.copy(source_path, destination_path)
    node = Node(name, public_key, private_key, random.randint(100, 1000), node_socket)
    nodes.append(node)
    node_public_key[i + 1] = public_key
    node_private_key[i + 1] = private_key
    if not re:
        if len(blockchain.chain) == 0:
            blockchain.create_genesis_block()
        initial_block = blockchain.chain[0]
        folder_path = "./blockchain/{}".format(name)
        os.makedirs(folder_path, exist_ok=True)
        file_path = os.path.join(folder_path, "block 0.json")
        with open(file_path, "w") as file:
            json.dump(initial_block.__dict__, file)
    if not re or flag == 1:
        if os.path.isfile("./blockchain/public_keys.json"):
            os.chmod("./blockchain/public_keys.json", 0o644)
        with open("./blockchain/public_keys.json", "w") as file:
            json.dump(node_public_key, file)
        if os.path.isfile("./blockchain/private_keys.json"):
            os.chmod("./blockchain/private_keys.json", 0o644)
        with open("./blockchain/private_keys.json", "w") as file:
            json.dump(node_private_key, file)
